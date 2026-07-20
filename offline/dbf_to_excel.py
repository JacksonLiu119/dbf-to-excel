import argparse
import sys
from pathlib import Path

from dbfread import DBF
from openpyxl import Workbook


EXCEL_MAX_ROWS = 1_048_576


def convert_dbf(dbf_path: Path, output_dir: Path, encoding: str) -> Path:
    table = DBF(
        str(dbf_path),
        encoding=encoding,
        char_decode_errors="replace",
        load=False,
    )

    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"{dbf_path.stem}.xlsx"

    workbook = Workbook(write_only=True)
    sheet_index = 1
    current_sheet = workbook.create_sheet(title=f"DBF_{sheet_index}")
    headers = list(table.field_names)
    current_sheet.append(headers)
    current_row_count = 1
    total_records = 0

    for record in table:
        if current_row_count >= EXCEL_MAX_ROWS:
            sheet_index += 1
            current_sheet = workbook.create_sheet(title=f"DBF_{sheet_index}")
            current_sheet.append(headers)
            current_row_count = 1

        current_sheet.append([record.get(field) for field in headers])
        current_row_count += 1
        total_records += 1

    workbook.save(output_path)
    print(f"OK  {dbf_path.name} -> {output_path.name} ({total_records} rows)")
    return output_path


def find_dbf_files(input_path: Path) -> list[Path]:
    if input_path.is_file():
        if input_path.suffix.lower() != ".dbf":
            raise ValueError(f"Not a DBF file: {input_path}")
        return [input_path]

    if input_path.is_dir():
        return sorted(input_path.glob("*.dbf"))

    raise FileNotFoundError(f"Path not found: {input_path}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Convert DBF files to Excel .xlsx files.",
    )
    parser.add_argument(
        "input",
        nargs="?",
        default="input",
        help="DBF file or folder. Default: input",
    )
    parser.add_argument(
        "-o",
        "--output",
        default="output",
        help="Output folder. Default: output",
    )
    parser.add_argument(
        "-e",
        "--encoding",
        default="cp950",
        help="DBF text encoding. Common values: cp950, big5, utf-8. Default: cp950",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    input_path = Path(args.input)
    output_dir = Path(args.output)

    try:
        dbf_files = find_dbf_files(input_path)
        if not dbf_files:
            print(f"No .dbf files found: {input_path}")
            return 1

        print(f"Found {len(dbf_files)} DBF file(s), encoding: {args.encoding}")
        for dbf_file in dbf_files:
            convert_dbf(dbf_file, output_dir, args.encoding)

        print("All files converted.")
        return 0
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
